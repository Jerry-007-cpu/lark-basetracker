"""Provider-independent table reading, filtering, snapshot, and diff logic."""

from __future__ import annotations

import csv
import hashlib
import json
import os
from collections import Counter
from datetime import date, datetime, timedelta
from typing import Any, Iterable


TITLE_KEYWORDS = [
    "名称", "标题", "主题", "事项", "项目", "记录", "岗位名称", "职位名称",
    "岗位", "职位", "职务", "岗位标题", "title", "name", "subject",
    "position", "job", "role",
]
DATE_KEYWORDS = [
    "更新时间", "最后更新时间", "修改时间", "更新日期", "变更时间",
    "发布时间", "发布日期", "创建时间", "创建日期", "开放时间", "开放日期",
    "上线时间", "上线日期", "开始时间", "截止时间", "截止日期", "投递开始",
    "投递截止", "时间", "日期", "date", "time", "updated", "modified",
    "created", "open", "publish",
]
KEY_KEYWORDS = [
    "唯一id", "记录id", "职位id", "岗位id", "编号", "序号", "投递链接",
    "职位链接", "岗位链接", "链接", "url", "link", "id",
]
DATE_TYPES = {5, 1001, 1002}
TEXTLIKE_TYPES = {1, 3, 15, 20}
STATE_SCHEMA_VERSION = 1


def expand(path: str) -> str:
    return os.path.abspath(os.path.expanduser(path))


def norm(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (date, datetime)):
        return value.isoformat(sep=" ") if isinstance(value, datetime) else value.isoformat()
    if isinstance(value, list):
        return ", ".join(part for part in (norm(item) for item in value) if part)
    if isinstance(value, dict):
        for key in ("text", "name", "en_name", "link", "url", "value"):
            if value.get(key) not in (None, ""):
                return norm(value[key])
        return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)
    return str(value)


def json_safe(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (date, datetime)):
        return value.isoformat(sep=" ") if isinstance(value, datetime) else value.isoformat()
    if isinstance(value, list):
        return [json_safe(item) for item in value]
    if isinstance(value, dict):
        return {str(key): json_safe(item) for key, item in value.items()}
    return str(value)


def keyword_score(name: str, keywords: Iterable[str]) -> int:
    lowered = name.lower().strip()
    best = 0
    for index, keyword in enumerate(keywords):
        candidate = keyword.lower()
        if lowered == candidate:
            best = max(best, 100 - index)
        elif candidate in lowered:
            best = max(best, 60 - index)
    return best


def auto_match_fields(fields_meta: list[dict[str, Any]]) -> dict[str, Any]:
    names = [str(field.get("field_name", "")) for field in fields_meta]
    title = next(
        (str(field.get("field_name")) for field in fields_meta if field.get("is_primary")),
        None,
    )
    if not title:
        scored = []
        for field in fields_meta:
            name = str(field.get("field_name", ""))
            score = keyword_score(name, TITLE_KEYWORDS)
            if field.get("type") in TEXTLIKE_TYPES:
                score += 5
            if score:
                scored.append((score, name))
        if scored:
            title = max(scored)[1]
    if not title:
        title = next(
            (str(field.get("field_name")) for field in fields_meta if field.get("type") in TEXTLIKE_TYPES),
            names[0] if names else None,
        )

    date_scored = []
    for field in fields_meta:
        name = str(field.get("field_name", ""))
        field_type = field.get("type")
        score = keyword_score(name, DATE_KEYWORDS)
        if field_type in DATE_TYPES:
            score += 40
        elif field_type == 2:
            score += 2
        if score:
            date_scored.append((score, name))
    date_scored.sort(reverse=True)

    key_scored = [(keyword_score(name, KEY_KEYWORDS), name) for name in names]
    key_scored = [item for item in key_scored if item[0] > 0]
    key_scored.sort(reverse=True)
    return {
        "title": title,
        "date": date_scored[0][1] if date_scored else None,
        "date_candidates": [name for _, name in date_scored],
        "key": key_scored[0][1] if key_scored else None,
        "all_names": names,
    }


def fields_meta_from_names(names: list[str]) -> list[dict[str, Any]]:
    return [
        {"field_name": name, "type": 1, "is_primary": False}
        for name in names
    ]


def to_epoch_ms(value: Any) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, list) and value:
        value = value[0]
    if isinstance(value, dict):
        value = value.get("value") or value.get("text") or ""
    if isinstance(value, datetime):
        return int(value.timestamp() * 1000)
    if isinstance(value, date):
        return int(datetime.combine(value, datetime.min.time()).timestamp() * 1000)
    if isinstance(value, (int, float)):
        number = int(value)
        return number * 1000 if 1_000_000_000 <= number < 10_000_000_000 else number
    text = str(value).strip()
    if text.isdigit():
        number = int(text)
        return number * 1000 if len(text) == 10 else number
    normalized = text.replace("Z", "+00:00")
    try:
        return int(datetime.fromisoformat(normalized).timestamp() * 1000)
    except ValueError:
        pass
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return int(datetime.strptime(text, fmt).timestamp() * 1000)
        except ValueError:
            continue
    return None


def resolve_range(days: int | None = None, since: str | None = None, until: str | None = None) -> tuple[int | None, int | None]:
    since_ms = until_ms = None
    if days is not None:
        since_ms = int((datetime.now() - timedelta(days=days)).timestamp() * 1000)
    if since:
        since_ms = to_epoch_ms(since)
        if since_ms is None:
            raise ValueError(f"无法解析起始日期：{since}")
    if until:
        end = datetime.strptime(until, "%Y-%m-%d") + timedelta(days=1)
        until_ms = int(end.timestamp() * 1000)
    return since_ms, until_ms


def filter_records(
    records: list[dict[str, Any]],
    date_field: str = "",
    days: int | None = None,
    since: str | None = None,
    until: str | None = None,
) -> tuple[list[tuple[int | None, dict[str, Any]]], int | None, int | None]:
    since_ms, until_ms = resolve_range(days=days, since=since, until=until)
    kept = []
    for record in records:
        fields = record.get("fields", record)
        date_ms = to_epoch_ms(fields.get(date_field)) if date_field else None
        if date_field and date_ms is None:
            continue
        if since_ms is not None and (date_ms is None or date_ms < since_ms):
            continue
        if until_ms is not None and (date_ms is None or date_ms >= until_ms):
            continue
        kept.append((date_ms, fields))
    kept.sort(key=lambda item: (item[0] is None, -(item[0] or 0)))
    return kept, since_ms, until_ms


def ms_to_date(milliseconds: int | None) -> str:
    if milliseconds is None:
        return ""
    return datetime.fromtimestamp(milliseconds / 1000).strftime("%Y-%m-%d")


def render_records(
    kept: list[tuple[int | None, dict[str, Any]]],
    title_field: str = "",
    date_field: str = "",
    show_fields: list[str] | None = None,
    since_ms: int | None = None,
    until_ms: int | None = None,
    heading: str = "表格更新整理",
) -> str:
    range_text = ""
    if since_ms is not None or until_ms is not None:
        start = ms_to_date(since_ms) if since_ms is not None else "…"
        end = ms_to_date(until_ms - 1) if until_ms is not None else "今天"
        range_text = f"（{start} ~ {end}）"
    lines = [f"📌 {heading}{range_text}　共 {len(kept)} 条"]
    if not kept:
        return "\n".join(lines + ["（该时间段内没有符合条件的记录）"])
    for date_ms, fields in kept:
        title = norm(fields.get(title_field)) if title_field else ""
        line = f"\n• {title or '(未命名记录)'}"
        if date_ms is not None:
            line += f"　日期：{ms_to_date(date_ms)}"
        lines.append(line)
        keys = show_fields or [key for key in fields if key not in (title_field, date_field)]
        for key in keys:
            value = norm(fields.get(key))
            if value:
                lines.append(f"    {key}：{value}")
    return "\n".join(lines)


def read_delimited_snapshot(file_path: str) -> tuple[list[str], list[dict[str, Any]]]:
    with open(file_path, "r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t")
        except csv.Error:
            dialect = csv.excel_tab if file_path.lower().endswith(".tsv") else csv.excel
        reader = csv.DictReader(handle, dialect=dialect)
        if not reader.fieldnames:
            raise ValueError("表格快照没有表头，无法识别字段。")
        names = [str(name).strip() for name in reader.fieldnames if name]
        rows = []
        for row in reader:
            rows.append({str(key).strip(): value for key, value in row.items() if key is not None})
    return names, rows


def read_xlsx_snapshot(file_path: str, sheet_name: str | None = None) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("读取 XLSX 需要 openpyxl；请安装后重试，或改为导出 CSV。") from exc
    workbook = load_workbook(file_path, read_only=False, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"找不到工作表“{sheet_name}”；可用工作表：{', '.join(workbook.sheetnames)}")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
        values = sheet.iter_rows()
        try:
            header_cells = next(values)
        except StopIteration as exc:
            raise ValueError("XLSX 快照为空。") from exc
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in header_cells]
        if not any(headers):
            raise ValueError("XLSX 快照第一行没有表头，无法识别字段。")
        rows = []
        for raw_row in values:
            fields = {}
            for index, cell in enumerate(raw_row):
                if index >= len(headers) or not headers[index]:
                    continue
                value = cell.hyperlink.target if cell.hyperlink and cell.hyperlink.target else cell.value
                fields[headers[index]] = "" if value is None else json_safe(value)
            if any(norm(value) for value in fields.values()):
                rows.append(fields)
        return [header for header in headers if header], rows
    finally:
        workbook.close()


def read_snapshot(path: str, sheet_name: str | None = None) -> tuple[list[str], list[dict[str, Any]]]:
    file_path = expand(path)
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"找不到快照文件：{file_path}")
    lowered = file_path.lower()
    if lowered.endswith((".csv", ".tsv")):
        return read_delimited_snapshot(file_path)
    if lowered.endswith(".xlsx"):
        return read_xlsx_snapshot(file_path, sheet_name=sheet_name)
    raise ValueError("当前快照模式支持 CSV、TSV 和 XLSX。")


def _record_key(fields: dict[str, Any], source_id: str, key_field: str, title_field: str) -> str:
    if key_field:
        value = norm(fields.get(key_field))
        if value:
            return value
    if source_id:
        return source_id
    if title_field:
        value = norm(fields.get(title_field))
        if value:
            return value
    digest = hashlib.sha256(
        json.dumps(json_safe(fields), ensure_ascii=False, sort_keys=True).encode("utf-8")
    ).hexdigest()[:16]
    return f"row-{digest}"


def build_state(
    records: list[dict[str, Any]],
    source: str = "",
    key_field: str = "",
    title_field: str = "",
) -> dict[str, Any]:
    prepared = []
    raw_keys = []
    for record in records:
        fields = json_safe(record.get("fields", record))
        source_id = norm(record.get("source_id") or record.get("record_id"))
        raw_key = _record_key(fields, source_id, key_field, title_field)
        raw_keys.append(raw_key)
        prepared.append({"key": raw_key, "source_id": source_id, "fields": fields})
    totals = Counter(raw_keys)
    seen = Counter()
    for record in prepared:
        key = record["key"]
        seen[key] += 1
        if totals[key] > 1:
            record["key"] = f"{key}#{seen[key]}"
    return {
        "schema_version": STATE_SCHEMA_VERSION,
        "captured_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "source": source,
        "key_field": key_field,
        "title_field": title_field,
        "duplicate_keys": sorted(key for key, count in totals.items() if count > 1),
        "records": prepared,
    }


def save_state(path: str, state: dict[str, Any]) -> None:
    output = expand(path)
    os.makedirs(os.path.dirname(output), exist_ok=True)
    with open(output, "w", encoding="utf-8") as handle:
        json.dump(state, handle, ensure_ascii=False, indent=2, sort_keys=True)
        handle.write("\n")


def load_state(path: str) -> dict[str, Any]:
    with open(expand(path), "r", encoding="utf-8") as handle:
        state = json.load(handle)
    if state.get("schema_version") != STATE_SCHEMA_VERSION or not isinstance(state.get("records"), list):
        raise ValueError("不是受支持的 lark-basetracker 快照状态文件。")
    return state


def diff_states(
    before: dict[str, Any],
    after: dict[str, Any],
    ignore_fields: Iterable[str] = (),
) -> dict[str, Any]:
    ignored = {field.strip() for field in ignore_fields if field.strip()}
    old_map = {record["key"]: record for record in before.get("records", [])}
    new_map = {record["key"]: record for record in after.get("records", [])}
    added = [new_map[key] for key in new_map.keys() - old_map.keys()]
    removed = [old_map[key] for key in old_map.keys() - new_map.keys()]
    changed = []
    for key in old_map.keys() & new_map.keys():
        old_fields = old_map[key].get("fields", {})
        new_fields = new_map[key].get("fields", {})
        changes = {}
        for field in (old_fields.keys() | new_fields.keys()) - ignored:
            old_value = norm(old_fields.get(field))
            new_value = norm(new_fields.get(field))
            if old_value != new_value:
                changes[field] = {"before": old_value, "after": new_value}
        if changes:
            changed.append({"key": key, "before": old_map[key], "after": new_map[key], "changes": changes})
    added.sort(key=lambda record: record["key"])
    removed.sort(key=lambda record: record["key"])
    changed.sort(key=lambda record: record["key"])
    return {"added": added, "removed": removed, "changed": changed}


def render_diff(diff: dict[str, Any], title_field: str = "", show_unchanged_fields: list[str] | None = None) -> str:
    added = diff.get("added", [])
    removed = diff.get("removed", [])
    changed = diff.get("changed", [])
    lines = [f"📌 表格快照变化　新增 {len(added)} · 修改 {len(changed)} · 删除 {len(removed)}"]

    def label(record: dict[str, Any]) -> str:
        fields = record.get("fields", {})
        return norm(fields.get(title_field)) if title_field else record.get("key", "")

    if added:
        lines.append("\n新增记录")
        for record in added:
            lines.append(f"• {label(record) or record['key']}")
            fields = record.get("fields", {})
            keys = show_unchanged_fields if show_unchanged_fields is not None else [key for key in fields if key != title_field]
            for field in keys:
                value = norm(record.get("fields", {}).get(field))
                if value:
                    lines.append(f"    {field}：{value}")
    if changed:
        lines.append("\n修改记录")
        for record in changed:
            lines.append(f"• {label(record['after']) or record['key']}")
            for field, values in sorted(record["changes"].items()):
                before_value = values["before"] or "（空）"
                after_value = values["after"] or "（空）"
                lines.append(f"    {field}：{before_value} → {after_value}")
    if removed:
        lines.append("\n删除记录")
        for record in removed:
            lines.append(f"• {label(record) or record['key']}")
    if not (added or changed or removed):
        lines.append("（两次快照之间没有变化）")
    return "\n".join(lines)


def parse_markdown_table(text: str) -> tuple[list[str], list[dict[str, str]]]:
    lines = [line.strip() for line in text.splitlines()]
    for index in range(len(lines) - 1):
        header = lines[index]
        separator = lines[index + 1]
        if "|" not in header or "|" not in separator:
            continue
        separator_cells = [cell.strip() for cell in separator.strip("|").split("|")]
        if not separator_cells or not all(cell and set(cell) <= {"-", ":", " "} for cell in separator_cells):
            continue
        headers = [cell.strip() for cell in header.strip("|").split("|")]
        rows = []
        for line in lines[index + 2:]:
            if "|" not in line:
                break
            values = [cell.strip() for cell in line.strip("|").split("|")]
            if len(values) < len(headers):
                values += [""] * (len(headers) - len(values))
            rows.append(dict(zip(headers, values)))
        if rows:
            return headers, rows
    raise ValueError("腾讯文档返回内容中没有识别到结构化表格。")
